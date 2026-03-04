import io
import os
import httpx
import asyncio
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database import engine, get_db, Base
from models import Numero

Base.metadata.create_all(bind=engine)


async def keep_alive():
    await asyncio.sleep(60)
    while True:
        try:
            url = os.environ.get("RENDER_EXTERNAL_URL", "http://localhost:8000")
            async with httpx.AsyncClient() as client:
                await client.get(f"{url}/api/stats", timeout=10)
        except:
            pass
        await asyncio.sleep(600)


@asynccontextmanager
async def lifespan(app):
    asyncio.create_task(keep_alive())
    yield


app = FastAPI(title="Rifa Manager", lifespan=lifespan)
templates = Jinja2Templates(directory="templates")


def init_numeros(db: Session):
    count = db.query(func.count(Numero.numero)).scalar()
    if count == 0:
        db.bulk_insert_mappings(Numero, [{"numero": i, "vendido": False} for i in range(1000)])
        db.commit()


@app.get("/", response_class=HTMLResponse)
def index(request: Request, db: Session = Depends(get_db)):
    total = db.query(func.count(Numero.numero)).scalar()
    if total == 0:
        init_numeros(db)
    vendidos = db.query(func.count(Numero.numero)).filter(Numero.vendido == True).scalar()
    disponibles = total - vendidos
    recaudado = db.query(func.sum(Numero.monto)).filter(Numero.vendido == True).scalar() or 0.0
    numeros = db.query(Numero).order_by(Numero.numero).all()
    return templates.TemplateResponse("index.html", {
        "request": request,
        "numeros": numeros,
        "vendidos": vendidos,
        "disponibles": disponibles,
        "recaudado": recaudado,
    })


@app.get("/api/numeros")
def get_numeros(db: Session = Depends(get_db)):
    numeros = db.query(Numero).order_by(Numero.numero).all()
    return [
        {
            "numero": n.numero,
            "vendido": n.vendido,
            "nombre": n.nombre,
            "referencia": n.referencia,
            "monto": n.monto,
            "metodo_pago": n.metodo_pago,
            "fecha": n.fecha.isoformat() if n.fecha else None,
        }
        for n in numeros
    ]


@app.get("/api/buscar/{numero}")
def buscar_numero(numero: int, db: Session = Depends(get_db)):
    if numero < 0 or numero > 999:
        raise HTTPException(status_code=400, detail="Número fuera de rango (0-999)")
    n = db.query(Numero).filter(Numero.numero == numero).first()
    if not n:
        raise HTTPException(status_code=404, detail="Número no encontrado")
    return {
        "numero": n.numero,
        "vendido": n.vendido,
        "nombre": n.nombre,
        "referencia": n.referencia,
        "monto": n.monto,
        "metodo_pago": n.metodo_pago,
        "fecha": n.fecha.isoformat() if n.fecha else None,
    }


@app.get("/api/buscar-nombre")
def buscar_por_nombre(q: str, db: Session = Depends(get_db)):
    if not q or len(q.strip()) < 2:
        raise HTTPException(status_code=400, detail="Ingresá al menos 2 caracteres")
    resultados = db.query(Numero).filter(
        Numero.vendido == True,
        Numero.nombre.ilike(f"%{q.strip()}%")
    ).order_by(Numero.numero).all()
    return [
        {
            "numero": n.numero,
            "nombre": n.nombre,
            "referencia": n.referencia,
            "monto": n.monto,
            "metodo_pago": n.metodo_pago,
            "fecha": n.fecha.isoformat() if n.fecha else None,
        }
        for n in resultados
    ]


@app.post("/api/vender")
def vender_numero(
    numero: int = Form(...),
    nombre: str = Form(...),
    referencia: str = Form(""),
    monto: float = Form(...),
    metodo_pago: str = Form(...),
    db: Session = Depends(get_db),
):
    if numero < 0 or numero > 999:
        raise HTTPException(status_code=400, detail="Número fuera de rango (0-999)")
    n = db.query(Numero).filter(Numero.numero == numero).with_for_update().first()
    if not n:
        raise HTTPException(status_code=404, detail="Número no encontrado")
    if n.vendido:
        raise HTTPException(status_code=409, detail=f"El número {numero} ya fue vendido a {n.nombre}")
    n.vendido = True
    n.nombre = nombre.strip()
    n.referencia = referencia.strip()
    n.monto = monto
    n.metodo_pago = metodo_pago
    n.fecha = datetime.now()
    db.commit()
    return {"ok": True, "numero": numero}


@app.post("/api/liberar")
def liberar_numero(numero: int = Form(...), db: Session = Depends(get_db)):
    n = db.query(Numero).filter(Numero.numero == numero).with_for_update().first()
    if not n:
        raise HTTPException(status_code=404, detail="Número no encontrado")
    if not n.vendido:
        raise HTTPException(status_code=409, detail="El número ya está disponible")
    n.vendido = False
    n.nombre = None
    n.referencia = None
    n.monto = None
    n.metodo_pago = None
    n.fecha = None
    db.commit()
    return {"ok": True}


@app.get("/api/stats")
def stats(db: Session = Depends(get_db)):
    vendidos = db.query(func.count(Numero.numero)).filter(Numero.vendido == True).scalar()
    recaudado = db.query(func.sum(Numero.monto)).filter(Numero.vendido == True).scalar() or 0.0
    return {"vendidos": vendidos, "disponibles": 1000 - vendidos, "recaudado": recaudado}


@app.get("/descargar-excel")
def descargar_excel(db: Session = Depends(get_db)):
    numeros = db.query(Numero).filter(Numero.vendido == True).order_by(Numero.numero).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Rifa"

    headers = ["Número", "Nombre", "Referencia", "Monto", "Método Pago", "Fecha"]
    header_fill = PatternFill("solid", fgColor="1a472a")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, n in enumerate(numeros, 2):
        ws.cell(row=row_idx, column=1, value=n.numero)
        ws.cell(row=row_idx, column=2, value=n.nombre)
        ws.cell(row=row_idx, column=3, value=n.referencia)
        ws.cell(row=row_idx, column=4, value=n.monto)
        ws.cell(row=row_idx, column=5, value=n.metodo_pago)
        ws.cell(row=row_idx, column=6, value=n.fecha.strftime("%Y-%m-%d %H:%M") if n.fecha else "")
        if row_idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="e8f5e9")

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rifa_vendidos.xlsx"},
    )